import streamlit as st
import random
import os
import io
import time
import sqlite3
from datetime import datetime
import cv2  # For face and eye detection

# Optional OpenPyXL for loading excel files if present
try:
    from openpyxl import load_workbook
    from PIL import Image as PILImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Import for live camera feed
try:
    from streamlit_webrtc import webrtc_streamer, VideoTransformerBase, RTCConfiguration
    WEBRTC_AVAILABLE = True
except ImportError:
    WEBRTC_AVAILABLE = False
    class VideoTransformerBase:
        pass

# Optional PyMongo for MongoDB
try:
    import pymongo
    from pymongo import MongoClient
    PYMONGO_AVAILABLE = True
except ImportError:
    PYMONGO_AVAILABLE = False


# ---------------------------
# Database Fallback (MongoDB -> SQLite)
# ---------------------------
SQLITE_DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "quiz_system.db")

def init_sqlite_db():
    conn = sqlite3.connect(SQLITE_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS apti_test (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id TEXT,
            timestamp TEXT,
            category TEXT,
            test_no INTEGER,
            no_of_questions INTEGER,
            marks_achieved INTEGER,
            time_taken REAL,
            avg_test_accuracy REAL
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS face_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id TEXT,
            timestamp TEXT,
            violation TEXT
        )
    """)
    conn.commit()
    conn.close()

init_sqlite_db()

def is_mongodb_available():
    if not PYMONGO_AVAILABLE:
        return False
    try:
        client = MongoClient("mongodb://localhost:27017/", serverSelectionTimeoutMS=1000)
        client.admin.command('ping')
        return True
    except Exception:
        return False

USE_MONGODB = is_mongodb_available()

def db_connect():
    if USE_MONGODB:
        client = MongoClient("mongodb://localhost:27017/")
        return client['quiz_system']
    return None

def store_face_log(student_id, message):
    """Log proctoring violations in MongoDB or SQLite."""
    if USE_MONGODB:
        try:
            db = db_connect()
            collection = db["face_logs"]
            log_data = {
                "student_id": student_id,
                "timestamp": datetime.now(),
                "violation": message
            }
            collection.insert_one(log_data)
            return
        except Exception:
            pass

    try:
        conn = sqlite3.connect(SQLITE_DB_PATH)
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO face_logs (student_id, timestamp, violation) VALUES (?, ?, ?)",
            (student_id, datetime.now().isoformat(), message)
        )
        conn.commit()
        conn.close()
    except Exception:
        pass


def get_test_number(username, category):
    if USE_MONGODB:
        try:
            db = db_connect()
            collection = db["apti_test"]
            latest_test = list(collection.find({"student_id": username, "category": category})
                               .sort("timestamp", pymongo.DESCENDING).limit(1))
            if latest_test:
                return latest_test[0]["test_no"] + 1
            return 1
        except Exception:
            pass

    try:
        conn = sqlite3.connect(SQLITE_DB_PATH)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT MAX(test_no) FROM apti_test WHERE student_id = ? AND category = ?",
            (username, category)
        )
        row = cursor.fetchone()
        conn.close()
        if row and row[0] is not None:
            return row[0] + 1
        return 1
    except Exception:
        return 1


def get_test_wise_accuracy(username, category, test_no):
    if USE_MONGODB:
        try:
            db = db_connect()
            collection = db['apti_test']
            test_details = collection.find({"student_id": username, "category": category, "test_no": test_no})
            correct_answers = 0
            total_questions = 0
            for test in test_details:
                correct_answers += test.get('marks_achieved', 0)
                total_questions += test.get('no_of_questions', 0)
            return round((correct_answers / total_questions * 100), 2) if total_questions > 0 else 0
        except Exception:
            pass

    try:
        conn = sqlite3.connect(SQLITE_DB_PATH)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT SUM(marks_achieved), SUM(no_of_questions) FROM apti_test WHERE student_id = ? AND category = ? AND test_no = ?",
            (username, category, test_no)
        )
        row = cursor.fetchone()
        conn.close()
        if row and row[1] and row[1] > 0:
            return round((row[0] / row[1]) * 100, 2)
        return 0
    except Exception:
        return 0


def get_average_accuracy(username, category, current_accuracy=None):
    if USE_MONGODB:
        try:
            db = db_connect()
            collection = db['apti_test']
            test_details = list(collection.find({"student_id": username, "category": category}))
            total_accuracy = 0
            test_count = 0
            for test in test_details:
                test_accuracy = get_test_wise_accuracy(username, category, test['test_no'])
                total_accuracy += test_accuracy
                test_count += 1
            if current_accuracy is not None:
                total_accuracy += current_accuracy
            avg_test_accuracy = (total_accuracy / (test_count + 1)) if test_count > 0 else (current_accuracy or 0)
            return round(avg_test_accuracy, 2)
        except Exception:
            pass

    try:
        conn = sqlite3.connect(SQLITE_DB_PATH)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT marks_achieved, no_of_questions FROM apti_test WHERE student_id = ? AND category = ?",
            (username, category)
        )
        rows = cursor.fetchall()
        conn.close()
        total_acc = 0
        count = 0
        for r in rows:
            if r[1] > 0:
                total_acc += (r[0] / r[1]) * 100
                count += 1
        if current_accuracy is not None:
            total_acc += current_accuracy
            count += 1
        return round(total_acc / count, 2) if count > 0 else (current_accuracy or 0)
    except Exception:
        return current_accuracy or 0


def store_test_details(username, test_no, category, no_of_questions, marks_achieved, time_taken, avg_test_accuracy):
    if USE_MONGODB:
        try:
            db = db_connect()
            collection = db['apti_test']
            existing_test = collection.find_one({"student_id": username, "test_no": test_no, "category": category})
            if not existing_test:
                collection.insert_one({
                    "student_id": username,
                    "timestamp": datetime.now(),
                    "category": category,
                    "test_no": test_no,
                    "no_of_questions": no_of_questions,
                    "marks_achieved": marks_achieved,
                    "time_taken": time_taken,
                    "avg_test_accuracy": avg_test_accuracy
                })
                st.success("Test details saved to database.")
                return
        except Exception:
            pass

    try:
        conn = sqlite3.connect(SQLITE_DB_PATH)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id FROM apti_test WHERE student_id = ? AND test_no = ? AND category = ?",
            (username, test_no, category)
        )
        if not cursor.fetchone():
            cursor.execute(
                """INSERT INTO apti_test 
                (student_id, timestamp, category, test_no, no_of_questions, marks_achieved, time_taken, avg_test_accuracy)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (username, datetime.now().isoformat(), category, test_no, no_of_questions, marks_achieved, time_taken, avg_test_accuracy)
            )
            conn.commit()
            st.success("Test details stored successfully (Local DB).")
        conn.close()
    except Exception as e:
        st.error(f"Error inserting test data: {e}")


# ---------------------------
# Video Transformer with Proctoring
# ---------------------------
class VideoTransformer(VideoTransformerBase):
    def __init__(self):
        try:
            self.face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
            self.eye_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_eye.xml")
        except Exception:
            self.face_cascade = None
            self.eye_cascade = None

        self.no_face_warning_count = 0
        self.multiple_face_warning_count = 0
        self.eye_gaze_warning_count = 0
        self.last_no_face_warning_time = time.time()
        self.last_multiple_warning_time = time.time()
        self.last_eye_gaze_warning_time = time.time()
        self.test_terminated = False

        self.no_face_frames = 0
        self.multiple_face_frames = 0
        self.eye_gaze_frames = 0
        self.frame_threshold = 5

        self.warning_interval = 2
        self.warning_limit = 10
        self.proctoring_enabled = False
        self.student_id = None

    def transform(self, frame):
        if not self.proctoring_enabled or self.face_cascade is None:
            return frame.to_ndarray(format="bgr24")

        img = frame.to_ndarray(format="bgr24")
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        current_time = time.time()
        violation_message = None

        faces = self.face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5)

        if len(faces) == 0:
            self.no_face_frames += 1
        else:
            self.no_face_frames = 0

        if self.no_face_frames >= self.frame_threshold:
            if current_time - self.last_no_face_warning_time > self.warning_interval:
                self.no_face_warning_count += 1
                self.last_no_face_warning_time = current_time
                if self.student_id:
                    store_face_log(self.student_id, "No Face Detected!")
            violation_message = "No Face Detected!"

        if len(faces) > 1:
            self.multiple_face_frames += 1
        else:
            self.multiple_face_frames = 0

        if self.multiple_face_frames >= self.frame_threshold:
            if current_time - self.last_multiple_warning_time > self.warning_interval:
                self.multiple_face_warning_count += 1
                self.last_multiple_warning_time = current_time
                if self.student_id:
                    store_face_log(self.student_id, "Multiple Faces Detected!")
            violation_message = "Multiple Faces Detected!"

        for (x, y, w, h) in faces:
            cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)

        if violation_message:
            overlay = img.copy()
            cv2.rectangle(overlay, (0, 0), (img.shape[1], img.shape[0]), (0, 0, 255), -1)
            cv2.addWeighted(overlay, 0.4, img, 0.6, 0, img)
            font = cv2.FONT_HERSHEY_SIMPLEX
            cv2.putText(img, violation_message, (30, 50), font, 1.0, (255, 255, 255), 2, cv2.LINE_AA)

        if (self.no_face_warning_count >= self.warning_limit) or \
           (self.multiple_face_warning_count >= self.warning_limit) or \
           (self.eye_gaze_warning_count >= self.warning_limit):
            self.test_terminated = True

        return img


# ---------------------------
# Default High-Quality Question Bank (Fallback)
# ---------------------------
DEFAULT_QUESTIONS_BANK = {
    'General': [
        {
            'question_no': 1,
            'question_text': 'A train running at the speed of 66 km/hr crosses a pole in 9 seconds. What is the length of the train?',
            'image_data': None,
            'options': ['120 meters', '155 meters', '165 meters', '180 meters'],
            'labeled_options': {'A': '120 meters', 'B': '155 meters', 'C': '165 meters', 'D': '180 meters'},
            'correct_answer': 'C',
            'explanation': 'Speed = 66 * (5/18) m/s = 55/3 m/s. Length = Speed * Time = (55/3) * 9 = 165 meters.'
        },
        {
            'question_no': 2,
            'question_text': 'A sum of money at simple interest amounts to Rs. 815 in 3 years and to Rs. 854 in 4 years. The sum is:',
            'image_data': None,
            'options': ['Rs. 650', 'Rs. 690', 'Rs. 698', 'Rs. 700'],
            'labeled_options': {'A': 'Rs. 650', 'B': 'Rs. 690', 'C': 'Rs. 698', 'D': 'Rs. 700'},
            'correct_answer': 'C',
            'explanation': 'S.I. for 1 year = 854 - 815 = Rs. 39. S.I. for 3 years = 39 * 3 = Rs. 117. Principal = 815 - 117 = Rs. 698.'
        },
        {
            'question_no': 3,
            'question_text': 'If A is the brother of B; B is the sister of C; and C is the father of D, how D is related to A?',
            'image_data': None,
            'options': ['Nephew / Niece', 'Brother', 'Uncle', 'Father'],
            'labeled_options': {'A': 'Nephew / Niece', 'B': 'Brother', 'C': 'Uncle', 'D': 'Father'},
            'correct_answer': 'A',
            'explanation': 'A, B, C are siblings. Since C is father of D, D is the nephew or niece of A.'
        },
        {
            'question_no': 4,
            'question_text': 'Find the odd one out from the given series: 3, 5, 11, 14, 17, 21',
            'image_data': None,
            'options': ['14', '21', '11', '3'],
            'labeled_options': {'A': '14', 'B': '21', 'C': '11', 'D': '3'},
            'correct_answer': 'A',
            'explanation': 'All numbers except 14 are odd numbers. 14 is an even number.'
        },
        {
            'question_no': 5,
            'question_text': 'Choose the word which is most nearly OPPOSITE in meaning to ANOMALOUS:',
            'image_data': None,
            'options': ['Desirable', 'Normal', 'Connected', 'Capacious'],
            'labeled_options': {'A': 'Desirable', 'B': 'Normal', 'C': 'Connected', 'D': 'Capacious'},
            'correct_answer': 'B',
            'explanation': 'Anomalous means deviating from what is standard, normal, or expected. The opposite is Normal.'
        },
        {
            'question_no': 6,
            'question_text': 'A trader sells an item for Rs. 480 with a loss of 20%. At what price should he sell it to gain 20%?',
            'image_data': None,
            'options': ['Rs. 600', 'Rs. 720', 'Rs. 680', 'Rs. 700'],
            'labeled_options': {'A': 'Rs. 600', 'B': 'Rs. 720', 'C': 'Rs. 680', 'D': 'Rs. 700'},
            'correct_answer': 'B',
            'explanation': '80% of CP = 480 => CP = 600. For 20% gain, SP = 120% of 600 = Rs. 720.'
        },
        {
            'question_no': 7,
            'question_text': 'In how many ways can the letters of the word LEADER be arranged?',
            'image_data': None,
            'options': ['720', '360', '120', '240'],
            'labeled_options': {'A': '720', 'B': '360', 'C': '120', 'D': '240'},
            'correct_answer': 'B',
            'explanation': 'Word has 6 letters with E repeated 2 times. Total arrangements = 6! / 2! = 720 / 2 = 360.'
        },
        {
            'question_no': 8,
            'question_text': 'A and B together can complete a work in 12 days, while B alone can finish it in 30 days. In how many days can A alone finish it?',
            'image_data': None,
            'options': ['18 days', '20 days', '24 days', '25 days'],
            'labeled_options': {'A': '18 days', 'B': '20 days', 'C': '24 days', 'D': '25 days'},
            'correct_answer': 'B',
            'explanation': 'A work rate = (1/12) - (1/30) = (5 - 2)/60 = 3/60 = 1/20. So A alone takes 20 days.'
        },
        {
            'question_no': 9,
            'question_text': 'Select the missing term in the sequence: AZ, CX, EV, GT, ?',
            'image_data': None,
            'options': ['IR', 'KP', 'HS', 'JQ'],
            'labeled_options': {'A': 'IR', 'B': 'KP', 'C': 'HS', 'D': 'JQ'},
            'correct_answer': 'A',
            'explanation': 'First letters increase by +2 (A, C, E, G, I). Second letters decrease (Z, X, V, T, R). Next is IR.'
        },
        {
            'question_no': 10,
            'question_text': 'Average of 5 numbers is 27. If one number is excluded, average becomes 25. The excluded number is:',
            'image_data': None,
            'options': ['30', '35', '40', '45'],
            'labeled_options': {'A': '30', 'B': '35', 'C': '40', 'D': '45'},
            'correct_answer': 'B',
            'explanation': 'Sum of 5 numbers = 5 * 27 = 135. Sum of 4 numbers = 4 * 25 = 100. Excluded number = 135 - 100 = 35.'
        }
    ],
    'Technical': [
        {
            'question_no': 1,
            'question_text': 'Which data structure follows the Last-In-First-Out (LIFO) principle?',
            'image_data': None,
            'options': ['Queue', 'Stack', 'Linked List', 'Binary Tree'],
            'labeled_options': {'A': 'Queue', 'B': 'Stack', 'C': 'Linked List', 'D': 'Binary Tree'},
            'correct_answer': 'B',
            'explanation': 'A Stack operates on a LIFO (Last In First Out) order.'
        },
        {
            'question_no': 2,
            'question_text': 'What is the worst-case time complexity of QuickSort?',
            'image_data': None,
            'options': ['O(N log N)', 'O(N)', 'O(N^2)', 'O(1)'],
            'labeled_options': {'A': 'O(N log N)', 'B': 'O(N)', 'C': 'O(N^2)', 'D': 'O(1)'},
            'correct_answer': 'C',
            'explanation': 'QuickSort worst-case time complexity occurs when the pivot selection is poor, leading to O(N^2).'
        },
        {
            'question_no': 3,
            'question_text': 'In C++, which keyword is used to prevent a class from being inherited?',
            'image_data': None,
            'options': ['sealed', 'final', 'static', 'const'],
            'labeled_options': {'A': 'sealed', 'B': 'final', 'C': 'static', 'D': 'const'},
            'correct_answer': 'B',
            'explanation': 'The `final` specifier prevents a class from being derived or a virtual function from being overridden.'
        },
        {
            'question_no': 4,
            'question_text': 'Which of the following is NOT a feature of Object-Oriented Programming (OOP)?',
            'image_data': None,
            'options': ['Encapsulation', 'Polymorphism', 'Compilation', 'Inheritance'],
            'labeled_options': {'A': 'Encapsulation', 'B': 'Polymorphism', 'C': 'Compilation', 'D': 'Inheritance'},
            'correct_answer': 'C',
            'explanation': 'Compilation is a build step performed by compilers, not an OOP concept.'
        },
        {
            'question_no': 5,
            'question_text': 'What will be the output of printing bool([]) in Python?',
            'image_data': None,
            'options': ['True', 'False', 'None', 'Error'],
            'labeled_options': {'A': 'True', 'B': 'False', 'C': 'None', 'D': 'Error'},
            'correct_answer': 'B',
            'explanation': 'An empty list in Python evaluates to False in boolean context.'
        },
        {
            'question_no': 6,
            'question_text': 'What is the main function of Virtual Memory in Operating Systems?',
            'image_data': None,
            'options': ['Speed up CPU cache', 'Extend physical RAM using disk space', 'Direct Memory Access', 'Prevent malware'],
            'labeled_options': {'A': 'Speed up CPU cache', 'B': 'Extend physical RAM using disk space', 'C': 'Direct Memory Access', 'D': 'Prevent malware'},
            'correct_answer': 'B',
            'explanation': 'Virtual Memory allows execution of processes larger than physical RAM by paging to secondary storage.'
        },
        {
            'question_no': 7,
            'question_text': 'Which SQL clause is used to filter groups created by GROUP BY?',
            'image_data': None,
            'options': ['WHERE', 'HAVING', 'ORDER BY', 'FILTER'],
            'labeled_options': {'A': 'WHERE', 'B': 'HAVING', 'C': 'ORDER BY', 'D': 'FILTER'},
            'correct_answer': 'B',
            'explanation': 'The HAVING clause is used to filter aggregated data after GROUP BY, whereas WHERE filters individual rows before aggregation.'
        },
        {
            'question_no': 8,
            'question_text': 'Which layer of the OSI model is responsible for routing IP packets across networks?',
            'image_data': None,
            'options': ['Data Link Layer', 'Network Layer', 'Transport Layer', 'Session Layer'],
            'labeled_options': {'A': 'Data Link Layer', 'B': 'Network Layer', 'C': 'Transport Layer', 'D': 'Session Layer'},
            'correct_answer': 'B',
            'explanation': 'The Network Layer (Layer 3) handles IP addressing and packet routing.'
        },
        {
            'question_no': 9,
            'question_text': 'What does ACID stand for in Database Systems?',
            'image_data': None,
            'options': ['Atomicity, Consistency, Isolation, Durability', 'Access, Control, Index, Data', 'Array, Chain, Index, Document', 'Async, Concurrent, Isolated, Distributed'],
            'labeled_options': {'A': 'Atomicity, Consistency, Isolation, Durability', 'B': 'Access, Control, Index, Data', 'C': 'Array, Chain, Index, Document', 'D': 'Async, Concurrent, Isolated, Distributed'},
            'correct_answer': 'A',
            'explanation': 'ACID guarantees database transaction reliability.'
        },
        {
            'question_no': 10,
            'question_text': 'Which algorithm is used for finding the shortest path in a graph with non-negative edge weights?',
            'image_data': None,
            'options': ['Prim algorithm', 'Dijkstra algorithm', 'Kruskal algorithm', 'Floyd-Warshall algorithm'],
            'labeled_options': {'A': 'Prim algorithm', 'B': 'Dijkstra algorithm', 'C': 'Kruskal algorithm', 'D': 'Floyd-Warshall algorithm'},
            'correct_answer': 'B',
            'explanation': 'Dijkstra algorithm calculates single-source shortest paths in graphs with non-negative edge weights.'
        }
    ]
}


def load_questions(category):
    questions = []
    category_list = ["aptitude", "data-interpretation", "verbal-ability", "logical-reasoning", "verbal-reasoning", "non-verbal-reasoning"] if category == 'General' else ["c-programming", "cpp-programming", "c-sharp-programming", "java-programming"]

    if OPENPYXL_AVAILABLE:
        for subcategory in category_list:
            file_name = f"{subcategory}.xlsx"
            file_path = os.path.join(os.path.dirname(__file__), file_name)
            if not os.path.exists(file_path):
                continue
            try:
                wb = load_workbook(file_path)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=2):
                    question_no = row[0].value
                    question_text = row[1].value
                    options = row[2].value
                    answer = row[3].value
                    explanation = row[4].value
                    img_path = None

                    if question_no and question_text and options and answer:
                        options_list = options.splitlines() if subcategory == 'non-verbal-reasoning' else str(options).split(';')
                        options_list = [option.strip() for option in options_list]
                        labeled_options = {chr(65 + i): option for i, option in enumerate(options_list)}
                        correct_label = str(answer).strip()
                        for label, option in labeled_options.items():
                            if option.strip().lower() == str(answer).strip().lower():
                                correct_label = label
                                break

                        questions.append({
                            'question_no': question_no,
                            'question_text': question_text,
                            'image_data': img_path,
                            'options': options_list,
                            'labeled_options': labeled_options,
                            'correct_answer': correct_label,
                            'explanation': str(explanation).strip() if explanation else "No explanation available."
                        })
            except Exception:
                pass

    if not questions:
        questions = DEFAULT_QUESTIONS_BANK.get(category, DEFAULT_QUESTIONS_BANK['General'])

    return questions


def rerun_app():
    if hasattr(st, 'rerun'):
        st.rerun()
    elif hasattr(st, 'experimental_rerun'):
        st.experimental_rerun()


# ---------------------------
# High-Contrast UI & Custom CSS
# ---------------------------
st.markdown("""
<style>
    /* Global Container */
    [data-testid="stAppViewContainer"] {
        background: linear-gradient(135deg, #0b0f19 0%, #111827 50%, #0f172a 100%) !important;
        color: #f8fafc !important;
    }
    
    /* Universal Text Readability */
    h1, h2, h3, h4, h5, h6, p, label, span, div, .stMarkdown, .stMarkdown p {
        color: #f8fafc !important;
    }

    /* Form Input Fields & Prompts */
    .stTextInput input, .stTextArea textarea {
        background-color: #1e293b !important;
        color: #ffffff !important;
        border: 1px solid #475569 !important;
        border-radius: 10px !important;
        font-size: 1rem !important;
    }
    .stTextInput label, .stTextArea label, .stSelectbox label, .stRadio label {
        color: #ffffff !important;
        font-weight: 700 !important;
        font-size: 1.05rem !important;
    }
    ::placeholder {
        color: #94a3b8 !important;
        opacity: 1 !important;
    }

    /* Radio Button Labels & Text */
    div[data-testid="stRadio"] label, div[data-testid="stRadio"] div[role="radiogroup"] label span {
        color: #ffffff !important;
        font-size: 1.05rem !important;
        font-weight: 500 !important;
    }

    /* Alert / Notice Containers */
    div[data-testid="stAlert"] {
        background-color: rgba(30, 41, 59, 0.9) !important;
        border: 1px solid rgba(99, 102, 241, 0.4) !important;
        border-radius: 12px !important;
    }
    div[data-testid="stAlert"] * {
        color: #f8fafc !important;
    }

    /* Buttons */
    .stButton button {
        background: linear-gradient(135deg, #6366f1 0%, #4f46e5 100%) !important;
        color: #ffffff !important;
        border-radius: 10px !important;
        font-weight: 700 !important;
        font-size: 1rem !important;
        border: none !important;
        box-shadow: 0 4px 14px 0 rgba(99, 102, 241, 0.4) !important;
    }
</style>
""", unsafe_allow_html=True)


st.title("🧠 AptiQuiz - Proctored Aptitude Test")
st.caption("Practice General & Technical Aptitude with Live Proctoring & Performance Analytics.")

# Sidebar: Live Camera Feed & Navigation
st.sidebar.title("📹 Proctoring & Live Camera")

camera = None
if WEBRTC_AVAILABLE:
    RTC_CONFIGURATION = RTCConfiguration({"iceServers": [{"urls": ["stun:stun.l.google.com:19302"]}]})
    with st.sidebar:
        try:
            camera = webrtc_streamer(
                key="camera",
                video_transformer_factory=VideoTransformer,
                rtc_configuration=RTC_CONFIGURATION,
                async_processing=True,
                media_stream_constraints={"video": True, "audio": False}
            )
        except Exception:
            st.sidebar.warning("Camera stream initialization notice: Camera offline or permission blocked.")
else:
    st.sidebar.info("Camera proctoring module: Running in Standard Mode.")

with st.sidebar:
    if camera and hasattr(camera, "video_transformer") and camera.video_transformer is not None:
        st.markdown(f"⚠️ **No Face Warnings:** {camera.video_transformer.no_face_warning_count}")
        st.markdown(f"⚠️ **Multiple Face Warnings:** {camera.video_transformer.multiple_face_warning_count}")

    st.markdown("---")

    if "questions" in st.session_state and st.session_state.questions:
        st.markdown("### 📌 Question Map")
        num_questions = len(st.session_state.questions)
        cols_per_row = 5
        rows = (num_questions + cols_per_row - 1) // cols_per_row
        for row in range(rows):
            cols = st.columns(cols_per_row)
            for col_index in range(cols_per_row):
                question_index = row * cols_per_row + col_index
                if question_index < num_questions:
                    button_label = str(question_index + 1)
                    with cols[col_index]:
                        if st.button(button_label, key=f"qbutton_{question_index}"):
                            st.session_state.current_question = question_index
                            rerun_app()

st.markdown("---")

# Main Quiz Section
if "started" not in st.session_state or not st.session_state.started:
    with st.container():
        st.markdown("### 🎯 Setup Your Assessment")
        username = st.text_input("Enter your Candidate Name / ID:", key="user_id_input", placeholder="e.g. Rahul Sharma")
        category = st.radio("Choose Test Category:", ['General', 'Technical'], horizontal=True)

        if username:
            test_no = get_test_number(username, category)
            st.info(f"📋 **Test Session Number:** {test_no}")
        else:
            test_no = 1

        test_type = st.radio("Select Test Duration Mode:",
                             ["⚡ Quick Challenge (10 Questions)", "🏆 Full Test (30 Questions)"], horizontal=True)
        no_of_questions = 10 if "Quick Challenge" in test_type else 30

        if st.button("🚀 Start Quiz Test"):
            if not username.strip():
                st.error("Please enter your name/username to start the test.")
            else:
                st.session_state.started = True
                st.session_state.username = username
                st.session_state.category = category
                st.session_state.test_no = test_no
                st.session_state.no_of_questions = no_of_questions

                questions_pool = load_questions(category)
                random.shuffle(questions_pool)
                st.session_state.questions = questions_pool[:no_of_questions]
                st.session_state.current_question = 0
                st.session_state.user_answers = [None] * len(st.session_state.questions)
                st.session_state.start_time = time.time()
                st.session_state.test_submitted = False

                if camera and hasattr(camera, "video_transformer") and camera.video_transformer:
                    camera.video_transformer.proctoring_enabled = True
                    camera.video_transformer.student_id = st.session_state.username

                rerun_app()

# Active Test / Results Section
if st.session_state.get("started", False):
    if (st.session_state.get("test_submitted", False) or
        st.session_state.get("test_terminated", False)):

        if camera and hasattr(camera, "video_transformer") and camera.video_transformer:
            camera.video_transformer.proctoring_enabled = False

        st.markdown("---")
        end_time = time.time()
        time_taken = round(end_time - st.session_state.get("start_time", end_time), 2)
        score = 0
        for i, q in enumerate(st.session_state.questions):
            if i < len(st.session_state.user_answers) and st.session_state.user_answers[i] == q["correct_answer"]:
                score += 1

        st.header("🎉 Assessment Completed!")
        c1, c2, c3 = st.columns(3)
        with c1:
            st.metric("Final Score", f"{score} / {len(st.session_state.questions)}")
        with c2:
            st.metric("Time Elapsed", f"{time_taken} sec")
        with c3:
            acc = round((score / len(st.session_state.questions)) * 100, 1) if st.session_state.questions else 0
            st.metric("Accuracy", f"{acc}%")

        if st.session_state.get("test_terminated", False):
            st.error("⚠️ Test was flagged and auto-submitted due to proctoring violation limit.")

        st.markdown("### 💡 Answer Breakdown & Key Explanations")
        for i, q in enumerate(st.session_state.questions):
            user_ans = st.session_state.user_answers[i] if i < len(st.session_state.user_answers) else "Not Answered"
            correct_ans = q["correct_answer"]
            is_correct = (user_ans == correct_ans)

            st.markdown(f"**Q{i + 1}: {q['question_text']}**")
            if is_correct:
                st.success(f"✅ Your Answer: {user_ans} (Correct)")
            else:
                st.error(f"❌ Your Answer: {user_ans} | Correct Answer: {correct_ans}")
            st.caption(f"💡 Explanation: {q['explanation']}")
            st.markdown("---")

        current_accuracy = get_test_wise_accuracy(st.session_state.username, st.session_state.category, st.session_state.test_no)
        avg_test_accuracy = get_average_accuracy(st.session_state.username, st.session_state.category, current_accuracy)

        store_test_details(
            st.session_state.username,
            st.session_state.test_no,
            st.session_state.category,
            len(st.session_state.questions),
            score,
            time_taken,
            avg_test_accuracy
        )

        if st.button("🔄 Retake New Assessment"):
            for key in ["started", "username", "category", "test_no", "questions", "current_question", "user_answers", "start_time", "test_terminated", "test_submitted"]:
                if key in st.session_state:
                    del st.session_state[key]
            rerun_app()

    else:
        current_index = st.session_state.get("current_question", 0)
        total_q = len(st.session_state.questions)

        if total_q > 0 and current_index < total_q:
            question_data = st.session_state.questions[current_index]

            st.markdown(f"### Question {current_index + 1} of {total_q}")
            st.progress((current_index + 1) / total_q)

            st.markdown(f"#### ❓ {question_data['question_text']}")

            options_keys = list(question_data["labeled_options"].keys())
            existing_ans = st.session_state.user_answers[current_index]
            default_idx = options_keys.index(existing_ans) if existing_ans in options_keys else None

            user_choice = st.radio(
                "Select your response:",
                options_keys,
                index=default_idx,
                format_func=lambda x: f"({x}) {question_data['labeled_options'][x]}",
                key=f"radio_q_{current_index}"
            )
            if user_choice is not None:
                st.session_state.user_answers[current_index] = user_choice

            b_col1, b_col2, b_col3 = st.columns(3)
            with b_col1:
                if st.button("⬅️ Previous Question") and current_index > 0:
                    st.session_state.current_question -= 1
                    rerun_app()

            with b_col2:
                if st.button("Next Question ➡️") and current_index < total_q - 1:
                    st.session_state.current_question += 1
                    rerun_app()

            with b_col3:
                if st.button("🏁 Submit Test"):
                    st.session_state.test_submitted = True
                    rerun_app()
